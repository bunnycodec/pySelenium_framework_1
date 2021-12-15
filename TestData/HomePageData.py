import openpyxl


class HomePageData:

    @staticmethod
    def test_data(testCaseName):
        book = openpyxl.load_workbook("/Users/bunny_codec/Library/Mobile Documents/com~apple~CloudDocs/Py_Dev/Selenium_Pytest_Template/TestData/Book1.xlsx")
        sheet = book.active

        data = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testCaseName:
                for j in range(2, sheet.max_column + 1):
                    data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

                return [data]

